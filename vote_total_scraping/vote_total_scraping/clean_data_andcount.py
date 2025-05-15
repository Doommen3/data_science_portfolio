import openpyxl

def clean_party_column(excel_file):
    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # Find the column index for "Party" using the header row (assumed to be in row 1)
    party_col_index = None
    for cell in ws[1]:
        if cell.value and str(cell.value).strip().lower() == "party":
            party_col_index = cell.column  # this returns an integer in modern openpyxl
            break

    if party_col_index is None:
        print("Party column not found in the Excel file.")
        return

    unique_parties = set()

    # Process all rows starting from row 2 (skipping header)
    for row in ws.iter_rows(min_row=2, min_col=party_col_index, max_col=party_col_index):
        cell = row[0]
        if cell.value:
            # Remove all '.' characters from the cell value
            cleaned_value = cell.value.replace(".", "")
            cell.value = cleaned_value  # update the cell value in the workbook
            # Add the cleaned party name to the set (strip extra whitespace)
            unique_parties.add(cleaned_value.strip())

    # Save the changes back to the same Excel file (or choose a new filename)
    wb.save(excel_file)

    # Print the unique party names and their count
    print("Unique parties found:")
    for party in sorted(unique_parties):
        print(party)
    print("Total unique parties:", len(unique_parties))

if __name__ == "__main__":
    excel_file = "output.xlsx"  # path to your output Excel file
    clean_party_column(excel_file)
