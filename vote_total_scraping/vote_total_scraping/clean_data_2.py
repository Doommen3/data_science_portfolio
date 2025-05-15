import openpyxl


def clean_party_column(input_file, cleaned_file, unique_file):
    # Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # Find the column indices for "Party" and "Year" using the header row (assumed to be in row 1)
    party_col_index = None
    year_col_index = None
    header_row = ws[1]

    for cell in header_row:
        if cell.value:
            header_value = str(cell.value).strip().lower()
            if header_value == "party":
                party_col_index = cell.column  # openpyxl columns are 1-indexed
            elif header_value == "year":
                year_col_index = cell.column

    if party_col_index is None:
        print("Error: 'Party' column not found in the Excel file.")
        return
    if year_col_index is None:
        print("Error: 'Year' column not found in the Excel file.")
        return

    # Set to hold unique (party, year) pairs
    unique_parties = set()

    # Process each row starting from row 2 (skipping the header)
    for row in ws.iter_rows(min_row=2):
        # Adjust indices from 1-indexed to 0-indexed for Python lists
        party_cell = row[party_col_index - 1]
        year_cell = row[year_col_index - 1]

        if party_cell.value:
            # Remove all '.' characters from the party name
            cleaned_party = str(party_cell.value).replace(".", "")
            party_cell.value = cleaned_party  # update the cell in the workbook

            # Get the corresponding year (if available)
            year_value = year_cell.value if year_cell.value is not None else ""
            # Add the cleaned party and its year to the set (strip any extra whitespace)
            unique_parties.add((cleaned_party.strip(), year_value))

    # Save the changes to a new Excel file for the cleaned data
    wb.save(cleaned_file)
    print(f"Cleaned data saved to '{cleaned_file}'.")

    # Create a new workbook for the unique party names along with their year
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = "Unique Parties"
    # Write the header row
    new_ws.append(["Party", "Year"])

    # Write each unique (party, year) pair into the new workbook
    # Sorting helps keep the list organized
    for party, year in sorted(unique_parties):
        new_ws.append([party, year])

    # Save the unique parties workbook
    new_wb.save(unique_file)
    print(f"Unique party-year pairs saved to '{unique_file}'.")
    print("Unique party-year pairs found:")
    for party, year in sorted(unique_parties):
        print(f"{party} - {year}")
    print("Total unique pairs:", len(unique_parties))


if __name__ == "__main__":
    # Define the file paths (adjust these as needed)
    input_file = "output_3.xlsx"  # Your original Excel file
    cleaned_file = "cleaned_file_3.xlsx"  # File with the party column cleaned (no periods)
    unique_file = "unique_parties_3.xlsx"  # File with unique party-year pairs

    clean_party_column(input_file, cleaned_file, unique_file)
